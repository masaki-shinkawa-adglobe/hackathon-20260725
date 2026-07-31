import csv
import io
from dataclasses import dataclass
from pathlib import PurePath

from starlette.datastructures import UploadFile


MAX_UPLOAD_SIZE = 10 * 1024 * 1024


class FileValidationError(Exception):
    def __init__(self, status_code: int, message: str) -> None:
        self.status_code = status_code
        self.message = message
        super().__init__(message)


@dataclass(frozen=True)
class ProcessedFile:
    filename: str
    mime_type: str
    text: str | None = None
    document: bytes | None = None


_MIME_TYPES = {
    ".pdf": {"application/pdf"},
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    },
    ".csv": {"text/csv", "application/csv"},
    ".txt": {"text/plain"},
}


async def read_upload_with_limit(upload: UploadFile) -> bytes:
    data = bytearray()
    while True:
        remaining = MAX_UPLOAD_SIZE + 1 - len(data)
        chunk = await upload.read(min(64 * 1024, remaining))
        if not chunk:
            return bytes(data)
        data.extend(chunk)
        if len(data) > MAX_UPLOAD_SIZE:
            raise FileValidationError(413, "File exceeds the 10 MiB size limit")


def process_upload(*, filename: str | None, content_type: str | None, data: bytes) -> ProcessedFile:
    suffix = PurePath(filename or "").suffix.lower()
    if suffix not in _MIME_TYPES:
        raise FileValidationError(415, "Unsupported file type")
    normalized_content_type = (content_type or "").split(";", 1)[0].lower()
    if normalized_content_type and normalized_content_type not in _MIME_TYPES[suffix]:
        raise FileValidationError(415, "File MIME type does not match its extension")
    if len(data) > MAX_UPLOAD_SIZE:
        raise FileValidationError(413, "File exceeds the 10 MiB size limit")

    if suffix == ".pdf":
        return ProcessedFile(filename=filename or "upload.pdf", mime_type="application/pdf", document=data)
    if suffix == ".xlsx":
        return ProcessedFile(
            filename=filename or "upload.xlsx",
            mime_type="text/plain",
            text=_extract_xlsx(data),
        )
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as error:
        raise FileValidationError(422, "CSV and TXT files must be UTF-8 encoded") from error
    return ProcessedFile(filename=filename or f"upload{suffix}", mime_type="text/plain", text=text)


def _extract_xlsx(data: bytes) -> str:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        output = io.StringIO()
        writer = csv.writer(output)
        for worksheet in workbook.worksheets:
            output.write(f"# Sheet: {worksheet.title}\n")
            for row in worksheet.iter_rows(values_only=True):
                writer.writerow(["" if cell is None else str(cell) for cell in row])
        workbook.close()
        return output.getvalue()
    except Exception as error:
        raise FileValidationError(422, "Invalid XLSX file") from error
